import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import os

import sys
import clr
from System import *
from System.Collections.Generic import List

sys.path.append(r"C:\Windows\Microsoft.NET\assembly\GAC_MSIL\OpenTDv242\v4.0_24.2.0.0__65e6d95ed5c2e178")
sys.path.append(r"C:\Windows\Microsoft.NET\assembly\GAC_MSIL\OpenTDv242.CoSolver\v4.0_24.2.0.0__65e6d95ed5c2e178")
sys.path.append(r"C:\Windows\Microsoft.NET\assembly\GAC_64\OpenTDv242.Results\v4.0_24.2.0.0__b62f614be6a1e14a")
clr.AddReference("OpenTDv242")
clr.AddReference("OpenTDv242.CoSolver")
clr.AddReference("OpenTDv242.Results")
from OpenTDv242 import *
from OpenTDv242.CoSolver import *
from OpenTDv242.Results.Dataset import SaveFile, ItemIdentifierCollection, DataTypes, StandardDataSubtypes


""" REQUIRED INPUTS"""

sav_files = [
    r'C:\Users\Path\To\Sav\File1.sav',
    r'C:\Users\Path\To\Sav\File2.sav'
]

excel_file = r'C:\Users\Path\To\Results.xlsx'


""" ANALYSIS OPTIONS """

# Set to True to calculate min/max from quasi-steady state only (final orbits)
# Set to False to calculate min/max from entire mission (includes transients)
USE_QUASI_STEADY_STATE_ONLY = True

# Orbit period in seconds (only used if USE_QUASI_STEADY_STATE_ONLY = True)
ORBIT_PERIOD_SECONDS = 5580  # ~93 minutes for LEO SSO
NUM_FINAL_ORBITS = 2  # Number of final orbits to use for quasi-steady analysis


""" PLOTTING INPUTS """

GENERATE_PLOTS = True  # Set to False to skip plotting
PLOT_OUTPUT_DIR = r'C:\Users\Path\To\Plot\Folder'

# Define submodels and their node groupings
# Format: {submodel_name: {group_name: [node_ids]}}
SUBMODELS_TO_PLOT = {
    'FS_ARRAYS': {
        'FS_Array1': list(range(1, 36)),
        'FS_Array2': list(range(37, 72))
    }
}


# Dictionary to store all results: {submodel: {case_name: {min, max, delta_min, delta_max}}}
all_results = {}
all_submodels = set()
# Dictionary to store raw data: {case_name: {submodel: {node_name: [temps_over_time]}}}
raw_data = {}
times_data = {}

# Process each sav file
for sav in sav_files:
    case_name = os.path.basename(sav)
    print(f"\n{'='*80}")
    print(f"Processing: {case_name}")
    print(f"{'='*80}")
    
    # Connect to Sav File
    data = SaveFile(sav)
    
    # Get submodels and times
    submodels = list(data.GetThermalSubmodels())
    times = data.GetTimes().GetValues()
    times_data[case_name] = times
    
    # Determine analysis time window
    if USE_QUASI_STEADY_STATE_ONLY:
        total_time = times[len(times) - 1]
        analysis_start_time = total_time - (NUM_FINAL_ORBITS * ORBIT_PERIOD_SECONDS)
        print(f"Analyzing {len(submodels)} submodels over {len(times)} time steps")
        print(f"Using QUASI-STEADY STATE: {analysis_start_time:.1f}s to {total_time:.1f}s (final {NUM_FINAL_ORBITS} orbits)\n")
    else:
        analysis_start_time = 0  # Use all data
        print(f"Analyzing {len(submodels)} submodels over {len(times)} time steps")
        print(f"Using FULL MISSION: All data from start to end\n")
    
    raw_data[case_name] = {}
    
    for submodel in submodels:
        print(f"\nSubmodel: {submodel}")
        
        # Get all node IDs in this submodel
        node_ids = list(data.GetNodeIds(submodel))
        
        if len(node_ids) == 0:
            print(f"  No nodes found - SKIPPING")
            continue
        
        # Only add to all_submodels if it has nodes
        all_submodels.add(submodel)
        
        # Build list of node names as strings
        node_names = [f"{submodel}.T{node_id}" for node_id in node_ids]
        
        # Get temperature data
        temps = data.GetData(*node_names)
        temp_values = temps.GetValues(Units.SI)  # Kelvin
        
        # Store raw data for this submodel (ALL data, for plotting)
        raw_data[case_name][submodel] = {}
        for i, node_name in enumerate(node_names):
            # Convert to Celsius
            temps_celsius = [t - 273.15 if not float('nan') == t else None for t in temp_values[i]]
            raw_data[case_name][submodel][node_name] = temps_celsius
        
        # Find min and max within the analysis window
        all_temps_in_window = []
        for node_temps in temp_values:
            for i, temp in enumerate(node_temps):
                # Only include temps from analysis window
                if times[i] >= analysis_start_time and not float('nan') == temp:
                    all_temps_in_window.append(temp)
        
        if len(all_temps_in_window) == 0:
            print(f"  No valid temperature data in analysis window")
            continue
        
        min_temp = min(all_temps_in_window) - 273.15  # Convert to Celsius
        max_temp = max(all_temps_in_window) - 273.15
        
        # Initialize submodel dict if needed
        if submodel not in all_results:
            all_results[submodel] = {}
        
        # Store results for this case
        all_results[submodel][case_name] = {
            'min_temp': round(min_temp, 2),
            'max_temp': round(max_temp, 2),
            'num_nodes': len(temp_values)
        }
        
        print(f"  Nodes: {len(temp_values)}")
        print(f"  Min Temp: {min_temp:.2f} °C")
        print(f"  Max Temp: {max_temp:.2f} °C")

print("\n" + "="*80)
print("All cases processed.")


"""PLOTTING"""

if GENERATE_PLOTS:
    import matplotlib.pyplot as plt
    import matplotlib
    matplotlib.use('Agg')  # Non-interactive backend for batch plotting
    
    # Create output directory if it doesn't exist
    os.makedirs(PLOT_OUTPUT_DIR, exist_ok=True)
    
    print("\n" + "="*80)
    print("GENERATING GROUPED AVERAGED PLOTS")
    print("="*80)
    
    # Calculate averaged data for specified submodel groups
    grouped_data = {}  # {case_name: {submodel: {group_name: [avg_temps_over_time]}}}
    
    for sav in sav_files:
        case_name = os.path.basename(sav)
        grouped_data[case_name] = {}
        
        for submodel, groups in SUBMODELS_TO_PLOT.items():
            if case_name not in raw_data or submodel not in raw_data[case_name]:
                print(f"⚠️  Warning: {submodel} not found in {case_name}")
                continue
            
            times = times_data[case_name]
            node_data = raw_data[case_name][submodel]
            
            grouped_data[case_name][submodel] = {}
            
            # Process each group (e.g., Array1, Array2)
            for group_name, node_ids in groups.items():
                # Filter nodes that belong to this group
                group_nodes = {}
                for node_name, temps in node_data.items():
                    # Extract node ID from node_name (format: "ARRAYS.T123")
                    node_id = int(node_name.split('.T')[1])
                    if node_id in node_ids:
                        group_nodes[node_name] = temps
                
                if not group_nodes:
                    print(f"⚠️  Warning: No nodes found for {submodel}.{group_name} in {case_name}")
                    continue
                
                # Calculate average temperature at each time step for this group
                avg_temps = []
                for i in range(len(times)):
                    temps_at_this_time = []
                    for node_temps in group_nodes.values():
                        if i < len(node_temps) and node_temps[i] is not None:
                            temps_at_this_time.append(node_temps[i])
                    
                    if temps_at_this_time:
                        avg_temps.append(sum(temps_at_this_time) / len(temps_at_this_time))
                    else:
                        avg_temps.append(None)
                
                grouped_data[case_name][submodel][group_name] = avg_temps
                print(f"  Averaged {len(group_nodes)} nodes for {submodel}.{group_name} in {case_name}")
    
    # Create plots for each case and submodel group
    for submodel, groups in SUBMODELS_TO_PLOT.items():
        print(f"\nPlotting {submodel} grouped average temperatures...")
        
        for sav in sav_files:
            case_name = os.path.basename(sav)
            case_name_clean = case_name.replace('.sav', '')
            
            if case_name not in grouped_data or submodel not in grouped_data[case_name]:
                continue
            
            times = times_data[case_name]
            
            # Convert times to Python list for easier indexing
            times_list = list(times)
            
            # Create two plots for this case: full mission and final orbits
            # Each plot will have multiple lines (one per group)
            fig1, ax1 = plt.subplots(figsize=(12, 6))
            fig2, ax2 = plt.subplots(figsize=(12, 6))
            
            colors = ['#1f77b4', '#ff7f0e', '#2ca02c', '#d62728', '#9467bd', '#8c564b']
            
            for idx, (group_name, avg_temps) in enumerate(grouped_data[case_name][submodel].items()):
                color = colors[idx % len(colors)]
                
                # PLOT 1: Full mission
                valid_data_full = [(t, temp) for t, temp in zip(times_list, avg_temps) if temp is not None]
                if valid_data_full:
                    plot_times_full, plot_temps_full = zip(*valid_data_full)
                    ax1.plot(plot_times_full, plot_temps_full, linewidth=2, color=color, label=group_name)
                
                # PLOT 2: Final orbits (quasi-steady)
                total_time = times_list[-1]
                final_orbits_start_time = total_time - (NUM_FINAL_ORBITS * ORBIT_PERIOD_SECONDS)
                
                valid_data_final = [(t, temp) for t, temp in zip(times_list, avg_temps) 
                                   if t >= final_orbits_start_time and temp is not None]
                
                if valid_data_final:
                    plot_times_final, plot_temps_final = zip(*valid_data_final)
                    # Normalize time to start at 0
                    plot_times_normalized = [t - final_orbits_start_time for t in plot_times_final]
                    ax2.plot(plot_times_normalized, plot_temps_final, linewidth=2, color=color, label=group_name)
            
            # Format PLOT 1 (Full Mission)
            ax1.set_xlabel('Time (s)', fontsize=12, fontweight='bold')
            ax1.set_ylabel('Average Temperature (°C)', fontsize=12, fontweight='bold')
            ax1.set_title(f'{submodel} - {case_name_clean}\nAverage Temperature vs Time (Full Mission)', 
                        fontsize=14, fontweight='bold')
            ax1.grid(True, alpha=0.3, linestyle='--')
            ax1.legend(loc='best', fontsize=10)
            
            # Remove margins on PLOT 1
            ax1.set_xlim(left=0, right=times_list[-1])
            ax1.margins(x=0)
            
            # Save PLOT 1
            plot1_filename = os.path.join(PLOT_OUTPUT_DIR, f'{case_name_clean}_{submodel}_full.png')
            fig1.tight_layout()
            fig1.savefig(plot1_filename, dpi=200, bbox_inches='tight')
            plt.close(fig1)
            print(f"  ✅ Saved: {plot1_filename}")
            
            # Format PLOT 2 (Final Orbits)
            ax2.set_xlabel(f'Time in Final {NUM_FINAL_ORBITS} Orbits (s)', fontsize=12, fontweight='bold')
            ax2.set_ylabel('Average Temperature (°C)', fontsize=12, fontweight='bold')
            ax2.set_title(f'{submodel} - {case_name_clean}\nAverage Temperature vs Time (Final {NUM_FINAL_ORBITS} Orbits - Quasi-Steady State)', 
                        fontsize=14, fontweight='bold')
            ax2.grid(True, alpha=0.3, linestyle='--')
            ax2.legend(loc='best', fontsize=10)
            
            # Remove margins on PLOT 2
            ax2.set_xlim(left=0, right=NUM_FINAL_ORBITS * ORBIT_PERIOD_SECONDS)
            ax2.margins(x=0)
            
            # Save PLOT 2
            plot2_filename = os.path.join(PLOT_OUTPUT_DIR, f'{case_name_clean}_{submodel}_final_orbit.png')
            fig2.tight_layout()
            fig2.savefig(plot2_filename, dpi=200, bbox_inches='tight')
            plt.close(fig2)
            print(f"  ✅ Saved: {plot2_filename}")
    
    print(f"\n✅ All plots saved to: {PLOT_OUTPUT_DIR}")

print("\n" + "="*80)
print("Creating Excel file...")
print("="*80)


""" EXCEL RESULTS """

# Read existing op limits and op limits library before modifying workbook
existing_margins_op_limits = {}
existing_op_limits_library = {}

try:
    wb = openpyxl.load_workbook(excel_file)
    print(f"Loaded existing workbook: {excel_file}")
    
    # Read op limits from Margins sheet if it exists
    if "Margins" in wb.sheetnames:
        ws_margins_old = wb["Margins"]
        if ws_margins_old.max_row >= 3:
            for row in range(3, ws_margins_old.max_row + 1):
                submodel = ws_margins_old.cell(row=row, column=1).value
                op_min = ws_margins_old.cell(row=row, column=2).value
                op_max = ws_margins_old.cell(row=row, column=3).value
                if submodel:
                    existing_margins_op_limits[submodel] = {
                        'min': op_min if op_min not in [None, ''] else None,
                        'max': op_max if op_max not in [None, ''] else None
                    }
        print(f"  Preserved {len(existing_margins_op_limits)} op limits from Margins sheet")
    
    # Read op limits library from Op Limits sheet if it exists
    if "Op Limits" in wb.sheetnames:
        ws_op_limits_old = wb["Op Limits"]
        if ws_op_limits_old.max_row >= 2:
            # Read header to get limit set names
            header_row = 2
            limit_set_names = []
            for col in range(2, ws_op_limits_old.max_column + 1, 2):  # Every 2 columns (Min/Max pair)
                limit_name = ws_op_limits_old.cell(row=header_row, column=col).value
                if limit_name and limit_name.endswith(' Min'):
                    limit_set_names.append(limit_name.replace(' Min', ''))
            
            # Read data rows
            for row in range(3, ws_op_limits_old.max_row + 1):
                submodel = ws_op_limits_old.cell(row=row, column=1).value
                if submodel:
                    existing_op_limits_library[submodel] = {}
                    for i, limit_set in enumerate(limit_set_names):
                        min_col = 2 + (i * 2)
                        max_col = min_col + 1
                        op_min = ws_op_limits_old.cell(row=row, column=min_col).value
                        op_max = ws_op_limits_old.cell(row=row, column=max_col).value
                        existing_op_limits_library[submodel][limit_set] = {
                            'min': op_min if op_min not in [None, ''] else None,
                            'max': op_max if op_max not in [None, ''] else None
                        }
        print(f"  Preserved Op Limits library with {len(existing_op_limits_library)} submodels")
    
    # Remove all existing sheets
    for sheet_name in wb.sheetnames:
        del wb[sheet_name]
    
except PermissionError:
    print(f"\n⚠️  ERROR: Cannot access {excel_file}")
    print("    The file is currently open. Please close it and try again.")
    exit()
    
except FileNotFoundError:
    wb = openpyxl.Workbook()
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    print(f"Created new workbook: {excel_file}")

# ============================================================================
# CREATE OP LIMITS LIBRARY SHEET
# ============================================================================
ws_op_limits = wb.create_sheet("Op Limits", 0)

print(f"Creating Op Limits library sheet...")

# Sort submodels for consistent ordering
sorted_submodels = sorted(all_submodels)

# Determine how many limit set columns to create
# If existing library has limit sets, use those; otherwise create 5 empty pairs
if existing_op_limits_library:
    # Get all unique limit set names from existing data
    all_limit_sets = set()
    for submodel_limits in existing_op_limits_library.values():
        all_limit_sets.update(submodel_limits.keys())
    limit_sets = sorted(all_limit_sets)
else:
    # Create 5 empty limit set pairs for user to fill in
    limit_sets = ['Limit Set 1', 'Limit Set 2', 'Limit Set 3', 'Limit Set 4', 'Limit Set 5']

# ROW 1: Title and Instructions
ws_op_limits.append(['OPERATIONAL TEMPERATURE LIMITS LIBRARY - Edit column headers to name your limit sets'])
title_cell = ws_op_limits.cell(row=1, column=1)
title_cell.font = Font(bold=True, size=12, color="FFFFFF")
title_cell.fill = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid")
title_cell.alignment = Alignment(horizontal='center', vertical='center')
ws_op_limits.merge_cells(start_row=1, start_column=1, end_row=1, end_column=1 + len(limit_sets) * 2)

# ROW 2: Headers (editable limit set names)
header_row = ['Submodel']
for limit_set in limit_sets:
    header_row.extend([f'{limit_set} Min', f'{limit_set} Max'])
ws_op_limits.append(header_row)

# Format header
for cell in ws_op_limits[2]:
    cell.font = Font(bold=True, size=10, color="000000")
    cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")  # Orange - indicates user can edit
    cell.alignment = Alignment(horizontal='center', vertical='center')

# ROW 3+: Data rows for each submodel
for submodel in sorted_submodels:
    row_data = [submodel]
    
    for limit_set in limit_sets:
        if submodel in existing_op_limits_library and limit_set in existing_op_limits_library[submodel]:
            op_min = existing_op_limits_library[submodel][limit_set]['min']
            op_max = existing_op_limits_library[submodel][limit_set]['max']
        else:
            op_min = ''
            op_max = ''
        row_data.extend([op_min, op_max])
    
    ws_op_limits.append(row_data)

# Auto-adjust column widths
for col_num in range(1, ws_op_limits.max_column + 1):
    column_letter = get_column_letter(col_num)
    max_length = 0
    for cell in ws_op_limits[column_letter]:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = min(max_length + 2, 30)
    ws_op_limits.column_dimensions[column_letter].width = adjusted_width

# Freeze first column and first 2 rows
ws_op_limits.freeze_panes = 'B3'

# ============================================================================
# CREATE MARGINS SHEET
# ============================================================================
ws_margins = wb.create_sheet("Margins", 1)

# ROW 1: Case filenames (merged across 4 columns each)
row1_data = ['Submodel', 'Op Min (°C)', 'Op Max (°C)']
for case_name in sav_files:
    row1_data.extend([os.path.basename(case_name), '', '', ''])  # Will merge these
ws_margins.append(row1_data)

# ROW 2: Column headers
row2_data = ['(Copy from Op Limits sheet)', '', '']  # Instruction under Submodel, Op Min, Op Max
for _ in sav_files:
    row2_data.extend(['Min', 'Max', 'ΔMin', 'ΔMax'])
ws_margins.append(row2_data)

# Format Row 1 (filenames)
for cell in ws_margins[1]:
    cell.font = Font(bold=True, size=11, color="FFFFFF")
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Format Row 2 (headers)
for cell in ws_margins[2]:
    cell.font = Font(bold=True, size=10, color="FFFFFF")
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Merge cells in Row 1 for each case filename
col_idx = 4  # Start after "Submodel", "Op Min", "Op Max"
for case_name in sav_files:
    ws_margins.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=col_idx+3)
    col_idx += 4

# ROW 3+: Data rows for each submodel
for submodel in sorted_submodels:
    row_data = [submodel]
    
    # Add operational limits (preserve from previous run)
    if submodel in existing_margins_op_limits:
        op_min = existing_margins_op_limits[submodel]['min']
        op_max = existing_margins_op_limits[submodel]['max']
    else:
        op_min = ''
        op_max = ''
    row_data.extend([op_min, op_max])
    
    # Add data for each case
    for sav in sav_files:
        case_name = os.path.basename(sav)
        
        if case_name in all_results.get(submodel, {}):
            case_data = all_results[submodel][case_name]
            min_temp = case_data['min_temp']
            max_temp = case_data['max_temp']
            
            # Calculate deltas
            delta_min = ''
            delta_max = ''
            if op_min not in ['', None]:
                delta_min = round(min_temp - float(op_min), 2)
            if op_max not in ['', None]:
                delta_max = round(float(op_max) - max_temp, 2)
            
            row_data.extend([min_temp, max_temp, delta_min, delta_max])
        else:
            # No data for this submodel in this case
            row_data.extend(['', '', '', ''])
    
    ws_margins.append(row_data)
    
    # Apply conditional formatting to delta columns
    row_num = ws_margins.max_row
    col_idx = 6  # First ΔMin column (after Submodel, Op Min, Op Max, Min, Max)
    
    for case_idx in range(len(sav_files)):
        delta_min_col = col_idx + (case_idx * 4)
        delta_max_col = delta_min_col + 1
        
        # Color ΔMin
        delta_min_cell = ws_margins.cell(row=row_num, column=delta_min_col)
        if delta_min_cell.value not in ['', None]:
            val = float(delta_min_cell.value)
            if val < 0:
                delta_min_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            elif val < 11:
                delta_min_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            else:
                delta_min_cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        
        # Color ΔMax
        delta_max_cell = ws_margins.cell(row=row_num, column=delta_max_col)
        if delta_max_cell.value not in ['', None]:
            val = float(delta_max_cell.value)
            if val < 0:
                delta_max_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            elif val < 11:
                delta_max_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            else:
                delta_max_cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

# Auto-adjust column widths for margins
for col_num in range(1, ws_margins.max_column + 1):
    column_letter = get_column_letter(col_num)
    max_length = 0
    for cell in ws_margins[column_letter]:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = min(max_length + 2, 50)  # Cap at 50
    ws_margins.column_dimensions[column_letter].width = adjusted_width

# Freeze first 3 columns (Submodel, Op Min, Op Max) and first 2 rows (headers)
ws_margins.freeze_panes = 'D3'

# ============================================================================
# CREATE RAW DATA SHEET (ALL CASES IN ONE TAB)
# ============================================================================
ws_raw = wb.create_sheet("Raw Data", 2)

print(f"Creating Raw Data sheet with all cases...")

current_row = 1

for sav in sav_files:
    case_name = os.path.basename(sav)
    times = times_data[case_name]
    
    # Add case filename row (merged across all columns)
    case_row = [case_name]
    ws_raw.append(case_row)
    case_cell = ws_raw.cell(row=current_row, column=1)
    case_cell.font = Font(bold=True, size=12, color="FFFFFF")
    case_cell.fill = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid")
    case_cell.alignment = Alignment(horizontal='center', vertical='center')
    # Merge across columns (Submodel, Node, + all time columns)
    num_cols = 2 + len(times)
    ws_raw.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=num_cols)
    current_row += 1
    
    # Header row: Submodel, Node, then all time values as numbers
    header_row = ['Submodel', 'Node'] + [round(t, 2) for t in times]
    ws_raw.append(header_row)
    for cell in ws_raw[current_row]:
        cell.font = Font(bold=True, size=10, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
    current_row += 1
    
    # Add data rows for each submodel and node
    for submodel in sorted_submodels:
        if case_name in raw_data and submodel in raw_data[case_name]:
            for node_name, temps in raw_data[case_name][submodel].items():
                row_data = [submodel, node_name] + [round(t, 2) if t is not None else '' for t in temps]
                ws_raw.append(row_data)
                current_row += 1
    
    # Add blank row between cases for readability
    ws_raw.append([])
    current_row += 1

# Auto-adjust column widths for raw data
for col_num in range(1, min(ws_raw.max_column + 1, 50)):  # Limit to first 50 columns for performance
    column_letter = get_column_letter(col_num)
    max_length = 0
    for i, cell in enumerate(ws_raw[column_letter]):
        if i > 100:  # Only check first 100 rows for performance
            break
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = min(max_length + 2, 30)  # Cap at 30
    ws_raw.column_dimensions[column_letter].width = adjusted_width

# Freeze first 2 columns (Submodel and Node)
ws_raw.freeze_panes = 'C1'

# Save workbook
wb.save(excel_file)
print(f"\nResults saved to: {excel_file}")
print(f"Total submodels: {len(sorted_submodels)}")
print(f"Total cases: {len(sav_files)}")
print(f"Sheets created: Op Limits + Margins + Raw Data")
print("\n✅ DONE! Open the Excel file to see all sheets.")
